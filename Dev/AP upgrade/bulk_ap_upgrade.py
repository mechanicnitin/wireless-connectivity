#!/usr/bin/env python3
"""
Bulk AP upgrade (immediate) driven by Excel/CSV sheet.

A) Pre-check skip if already on target version (per scope + allowed models)
B) Model targeting guard (default AP45; extend via .env ALLOWED_MODELS=AP45,AP47)
C) Rate limiting + retry/backoff for API reliability

Inputs:
- .env:
    MIST_BASE_URL, MIST_ORG_ID, MIST_ACCESS_TOKEN
    Proxy: ALL_PROXY or HTTPS_PROXY/HTTP_PROXY, and NO_PROXY

- Excel/CSV columns required:
    site_name, target_version, scope
  scope values:
    all | connected

Usage:
  python bulk_ap_upgrade.py -x site_list.xlsx --dry_run --preflight
  python bulk_ap_upgrade.py -x site_list.xlsx --yes
  python bulk_ap_upgrade.py -c site_list.csv --yes
"""

import sys
import os
import json
import getopt
import logging
import time
import csv
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook

LOG_FILE = "./script.log"
ENV_FILE = ".env"
INPUT_FILE: Optional[str] = None
SHEET_NAME: Optional[str] = None
DRY_RUN = False
DO_PREFLIGHT = False
AUTO_YES = False
INPUT_FORMAT = "xlsx"  # "xlsx" or "csv"

LOGGER = logging.getLogger(__name__)


class console:
    @staticmethod
    def info(msg: str):
        print(f"[INFO] {msg}")

    @staticmethod
    def warning(msg: str):
        print(f"[WARN] {msg}", file=sys.stderr)

    @staticmethod
    def error(msg: str):
        print(f"[ERROR] {msg}", file=sys.stderr)

    @staticmethod
    def critical(msg: str):
        print(f"[CRITICAL] {msg}", file=sys.stderr)


def load_env_file(path: str) -> Dict[str, str]:
    expanded = os.path.expanduser(path)
    if not os.path.exists(expanded):
        raise FileNotFoundError(f"Env file not found: {expanded}")

    env: Dict[str, str] = {}
    with open(expanded, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#") or "=" not in s:
                continue
            k, v = s.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def normalize_scope(scope: str) -> str:
    s = (scope or "").strip().lower()
    if s == "all":
        return "all"
    if s in ("connected", "connected_only", "online"):
        return "connected"
    raise ValueError(f"Invalid scope '{scope}'. Use 'all' or 'connected'.")


def read_csv_rows(path: str) -> List[Dict[str, str]]:
    """Read CSV file with required columns: site_name, target_version, scope"""
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError("CSV is empty")
        
        required = ["site_name", "target_version", "scope"]
        missing = [c for c in required if c not in reader.fieldnames]
        if missing:
            raise ValueError(f"Missing required column(s): {missing}. Found: {reader.fieldnames}")
        
        for row_num, row in enumerate(reader, start=2):
            site_name = (row.get("site_name") or "").strip()
            if not site_name:
                continue
            rows.append({
                "site_name": site_name,
                "target_version": (row.get("target_version") or "").strip(),
                "scope": (row.get("scope") or "").strip(),
            })
    return rows


def read_excel_rows(path: str, sheet_name: Optional[str]) -> List[Dict[str, str]]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {h.lower(): i for i, h in enumerate(headers) if h}

    required = ["site_name", "target_version", "scope"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(f"Missing required column(s): {missing}. Found headers: {headers}")

    out: List[Dict[str, str]] = []
    for r in rows[1:]:
        if r is None:
            continue
        site_name = r[idx["site_name"]]
        target_version = r[idx["target_version"]]
        scope = r[idx["scope"]]

        if site_name is None or str(site_name).strip() == "":
            continue

        out.append(
            {
                "site_name": str(site_name).strip(),
                "target_version": str(target_version).strip() if target_version is not None else "",
                "scope": str(scope).strip() if scope is not None else "",
            }
        )
    return out


def parse_allowed_models(env: Dict[str, str]) -> List[str]:
    raw = (env.get("ALLOWED_MODELS") or "AP45").strip()
    models = [m.strip() for m in raw.split(",") if m.strip()]
    return models if models else ["AP45"]


def parse_int(env: Dict[str, str], key: str, default: int) -> int:
    v = env.get(key)
    if v is None or str(v).strip() == "":
        return default
    try:
        return int(str(v).strip())
    except Exception:
        return default


def parse_float(env: Dict[str, str], key: str, default: float) -> float:
    v = env.get(key)
    if v is None or str(v).strip() == "":
        return default
    try:
        return float(str(v).strip())
    except Exception:
        return default


class MistClient:
    def __init__(
        self,
        base_url: str,
        token: str,
        timeout: int = 60,
        proxies: Optional[Dict[str, str]] = None,
        no_proxy: Optional[str] = None,
        max_retries: int = 5,
    ):
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self.max_retries = max_retries

        self.session = requests.Session()
        self.session.trust_env = False

        self.session.headers.update(
            {
                "Content-Type": "application/json",
                "Authorization": f"Token {token}",
                "Accept": "application/json",
            }
        )

        if proxies:
            self.session.proxies.update(proxies)

        if no_proxy:
            os.environ["NO_PROXY"] = no_proxy
            os.environ["no_proxy"] = no_proxy

    def _url(self, path: str) -> str:
        return f"{self.base_url}{path}"

    def _sleep_backoff(self, attempt: int, retry_after: Optional[str] = None):
        if retry_after:
            try:
                ra = int(retry_after)
                time.sleep(max(0, ra))
                return
            except Exception:
                pass
        delay = min(30, (2 ** attempt))
        time.sleep(delay)

    def request(self, method: str, path: str, params: Optional[Dict[str, Any]] = None, payload: Any = None) -> Any:
        data = json.dumps(payload) if payload is not None else None
        url = self._url(path)

        last_err: Optional[Exception] = None
        for attempt in range(0, self.max_retries + 1):
            try:
                resp = self.session.request(
                    method=method,
                    url=url,
                    params=params,
                    data=data,
                    timeout=self.timeout,
                )

                if resp.status_code in (429, 500, 502, 503, 504):
                    retry_after = resp.headers.get("Retry-After")
                    if attempt < self.max_retries:
                        console.warning(
                            f"{method} {path} -> {resp.status_code}. Retrying (attempt {attempt+1}/{self.max_retries})..."
                        )
                        self._sleep_backoff(attempt, retry_after=retry_after)
                        continue

                if not resp.ok:
                    raise RuntimeError(f"{method} {path} failed ({resp.status_code}): {resp.text}")

                return resp.json() if resp.text.strip() else None

            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                last_err = e
                if attempt < self.max_retries:
                    console.warning(
                        f"{method} {path} -> network error: {e}. Retrying (attempt {attempt+1}/{self.max_retries})..."
                    )
                    self._sleep_backoff(attempt)
                    continue
                raise
            except Exception as e:
                last_err = e
                raise

        if last_err:
            raise last_err
        raise RuntimeError(f"{method} {path} failed unexpectedly")

    def get_all_pages(self, path: str, params: Optional[Dict[str, Any]] = None, limit: int = 1000) -> List[Any]:
        items: List[Any] = []
        page = 1
        while True:
            p = dict(params or {})
            p["limit"] = limit
            p["page"] = page

            batch = self.request("GET", path, params=p)
            if not isinstance(batch, list):
                break

            items.extend(batch)
            if len(batch) < limit:
                break
            page += 1

        return items


def http_preflight(client: MistClient, org_id: str):
    console.info("Running HTTP preflight (via proxy if configured)...")
    client.request("GET", f"/orgs/{org_id}/sites", params={"limit": 1, "page": 1})
    console.info("HTTP preflight OK (connectivity/auth looks good).")


def build_site_name_to_id(client: MistClient, org_id: str) -> Dict[str, List[Tuple[str, str]]]:
    sites = client.get_all_pages(f"/orgs/{org_id}/sites", limit=1000)
    mapping: Dict[str, List[Tuple[str, str]]] = {}
    for s in sites:
        name = (s.get("name") or "").strip()
        sid = s.get("id")
        if not name or not sid:
            continue
        mapping.setdefault(name.casefold(), []).append((name, sid))
    return mapping


def resolve_site_id(site_name: str, mapping: Dict[str, List[Tuple[str, str]]]) -> Tuple[Optional[str], str]:
    matches = mapping.get(site_name.strip().casefold(), [])
    if len(matches) == 1:
        return matches[0][1], matches[0][0]
    if len(matches) > 1:
        return None, f"Ambiguous site_name '{site_name}' (matches: {[m[0] for m in matches]})"
    return None, f"Site not found for site_name '{site_name}'"


def get_site_devices_stats(client: MistClient, site_id: str) -> List[Dict[str, Any]]:
    data = client.request("GET", f"/sites/{site_id}/stats/devices")
    return data if isinstance(data, list) else []


def is_allowed_model(device: Dict[str, Any], allowed_models: List[str]) -> bool:
    model = (device.get("model") or "").strip().upper()
    if not model:
        return False
    return model in {m.strip().upper() for m in allowed_models}


def is_connected(device: Dict[str, Any]) -> bool:
    return (device.get("status") or "").strip().lower() == "connected"


def get_device_id(device: Dict[str, Any]) -> Optional[str]:
    return device.get("id") or device.get("device_id")


def get_device_version(device: Dict[str, Any]) -> Optional[str]:
    for key in ("version", "firmware_version", "firmware", "sw_version"):
        v = device.get(key)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def select_target_ap_device_ids(
    devices: List[Dict[str, Any]],
    scope: str,
    allowed_models: List[str],
) -> List[str]:
    out: List[str] = []
    for d in devices:
        if not is_allowed_model(d, allowed_models):
            continue
        if scope == "connected" and not is_connected(d):
            continue
        did = get_device_id(d)
        if did:
            out.append(did)
    return out


def precheck_already_on_target(
    devices: List[Dict[str, Any]],
    target_device_ids: List[str],
    target_version: str,
) -> Tuple[bool, int, int]:
    target_set = set(target_device_ids)
    known = 0
    on_target = 0
    total = 0

    for d in devices:
        did = get_device_id(d)
        if not did or did not in target_set:
            continue
        total += 1
        v = get_device_version(d)
        if v is None:
            continue
        known += 1
        if v == target_version:
            on_target += 1

    if total == 0:
        return False, known, total
    if known < total:
        return False, known, total
    return on_target == total, known, total


def upgrade_site(
    client: MistClient,
    site_id: str,
    version: str,
    scope: str,
    devices_stats: List[Dict[str, Any]],
    allowed_models: List[str],
) -> Dict[str, Any]:
    device_ids = select_target_ap_device_ids(devices_stats, scope, allowed_models)
    payload: Dict[str, Any] = {
        "version": version,
        "enable_p2p": False,
        "device_ids": device_ids,
    }
    return client.request("POST", f"/sites/{site_id}/devices/upgrade", payload=payload)


def usage():
    print(
        """
Bulk AP Upgrade from Excel/CSV (Immediate)

Required (one of):
  -x, --excel=        Excel file path (.xlsx)
  -c, --csv=          CSV file path (.csv)

Optional:
  -e, --env=          Env file path (default: ./.env)
  -l, --log_file=     Log file path (default: ./script.log)
  --sheet=            Sheet name (default: first sheet, Excel only)
  --dry_run           Don't call upgrade API; only validate & show plan
  --preflight         Do a quick API call to validate proxy/auth before processing
  --yes               Automatically proceed with upgrades (non-interactive)

File columns required:
  site_name, target_version, scope

scope values:
  all | connected

Env optional:
  ALLOWED_MODELS=AP45[,AP47]
  RATE_LIMIT_DELAY_SECONDS=1
  MAX_RETRIES=5

Example:
  python bulk_ap_upgrade.py -x site_list.xlsx --dry_run --preflight
  python bulk_ap_upgrade.py -c site_list.csv --yes
  python bulk_ap_upgrade.py -x site_list.xlsx --yes
"""
    )
    sys.exit(0)


if __name__ == "__main__":
    try:
        opts, _ = getopt.getopt(
            sys.argv[1:],
            "hx:c:e:l:",
            ["help", "excel=", "csv=", "env=", "log_file=", "sheet=", "dry_run", "preflight", "yes"],
        )
    except getopt.GetoptError as err:
        console.error(str(err))
        usage()

    for o, a in opts:
        if o in ("-h", "--help"):
            usage()
        elif o in ("-x", "--excel"):
            INPUT_FILE = a
            INPUT_FORMAT = "xlsx"
        elif o in ("-c", "--csv"):
            INPUT_FILE = a
            INPUT_FORMAT = "csv"
        elif o in ("-e", "--env"):
            ENV_FILE = a
        elif o in ("-l", "--log_file"):
            LOG_FILE = a
        elif o == "--sheet":
            SHEET_NAME = a
        elif o == "--dry_run":
            DRY_RUN = True
        elif o == "--preflight":
            DO_PREFLIGHT = True
        elif o == "--yes":
            AUTO_YES = True

    if not INPUT_FILE:
        console.critical("Input file is required (-x / --excel or -c / --csv).")
        usage()

    logging.basicConfig(filename=LOG_FILE, filemode="w")
    LOGGER.setLevel(logging.DEBUG)

    try:
        env = load_env_file(ENV_FILE)
    except Exception as e:
        console.critical(str(e))
        sys.exit(1)

    base_url = env.get("MIST_BASE_URL")
    org_id = env.get("MIST_ORG_ID")
    token = env.get("MIST_ACCESS_TOKEN")

    if not base_url or not org_id or not token:
        console.critical("Missing MIST_BASE_URL and/or MIST_ORG_ID and/or MIST_ACCESS_TOKEN in env file.")
        sys.exit(1)

    allowed_models = parse_allowed_models(env)
    rate_delay = parse_float(env, "RATE_LIMIT_DELAY_SECONDS", 1.0)
    max_retries = parse_int(env, "MAX_RETRIES", 5)

    proxy = env.get("ALL_PROXY") or env.get("HTTPS_PROXY") or env.get("HTTP_PROXY")
    proxies = {"http": proxy, "https": proxy} if proxy else None
    no_proxy = env.get("NO_PROXY")

    if proxies:
        console.info(f"Using proxy for http/https: {proxy}")
    if no_proxy:
        console.info(f"NO_PROXY set to: {no_proxy}")

    console.info(f"Allowed AP models: {', '.join(allowed_models)}")
    console.info(f"Rate limit delay per site: {rate_delay}s")
    console.info(f"HTTP max retries: {max_retries}")

    client = MistClient(
        base_url=base_url,
        token=token,
        proxies=proxies,
        no_proxy=no_proxy,
        max_retries=max_retries,
    )

    if DO_PREFLIGHT:
        try:
            http_preflight(client, org_id)
        except Exception as e:
            console.critical(f"Preflight failed: {e}")
            console.critical("Check proxy reachability, credentials, and whether proxy requires auth.")
            sys.exit(1)

    # Read input file
    try:
        if INPUT_FORMAT == "csv":
            rows = read_csv_rows(INPUT_FILE)
        else:
            rows = read_excel_rows(INPUT_FILE, SHEET_NAME)
    except Exception as e:
        console.critical(f"Failed to read input file: {e}")
        sys.exit(1)

    if not rows:
        console.critical("No usable rows found in input file.")
        sys.exit(1)

    # Build site lookup
    try:
        site_map = build_site_name_to_id(client, org_id)
    except Exception as e:
        console.critical(f"Failed to load sites from Mist. Error: {e}")
        sys.exit(1)

    plan: List[Dict[str, Any]] = []
    errors: List[str] = []

    for excel_row_num, r in enumerate(rows, start=2):
        site_name = r["site_name"]
        version = r["target_version"]
        scope_raw = r["scope"]

        if not version:
            errors.append(f"Row {excel_row_num}: target_version is empty for site '{site_name}'")
            continue

        try:
            scope = normalize_scope(scope_raw)
        except Exception as e:
            errors.append(f"Row {excel_row_num}: {e}")
            continue

        site_id, msg = resolve_site_id(site_name, site_map)
        if not site_id:
            errors.append(f"Row {excel_row_num}: {msg}")
            continue

        plan.append(
            {
                "row": excel_row_num,
                "site_name": msg,
                "site_id": site_id,
                "target_version": version,
                "scope": scope,
            }
        )

    console.info(f"Loaded {len(rows)} row(s) from input file.")
    console.info(f"Planned upgrades: {len(plan)} site(s). Errors: {len(errors)}.")

    if errors:
        console.warning("Some rows were skipped due to errors:")
        for e in errors:
            console.warning(f"  - {e}")

    if not plan:
        console.critical("No valid sites to upgrade after validation. Exiting.")
        sys.exit(1)

    for p in plan[:20]:
        print(f"Row {p['row']}: {p['site_name']} | scope={p['scope']} | version={p['target_version']}")
    if len(plan) > 20:
        print(f"... ({len(plan)-20} more)")

    if DRY_RUN:
        console.info("DRY RUN enabled. No upgrades were triggered.")
        sys.exit(0)

    if not AUTO_YES:
        resp = input(f"\nProceed to trigger upgrades for {len(plan)} site(s)? (y/N) ")
        if resp.strip().lower() != "y":
            console.info("Cancelled.")
            sys.exit(0)

    success = 0
    failed = 0
    skipped = 0

    for idx, p in enumerate(plan, start=1):
        site = p["site_name"]
        sid = p["site_id"]
        version = p["target_version"]
        scope = p["scope"]

        if idx > 1 and rate_delay > 0:
            time.sleep(rate_delay)

        try:
            devices_stats = get_site_devices_stats(client, sid)
            target_device_ids = select_target_ap_device_ids(devices_stats, scope, allowed_models)

            if not target_device_ids:
                console.warning(f"SKIP: {site} -> no eligible {allowed_models} APs found for scope='{scope}'")
                skipped += 1
                continue

            all_on_target, known, total = precheck_already_on_target(
                devices_stats, target_device_ids, version
            )
            if all_on_target:
                console.info(
                    f"SKIP: {site} -> already on target version {version} (checked {known}/{total})"
                )
                skipped += 1
                continue

            result = upgrade_site(client, sid, version, scope, devices_stats, allowed_models)
            console.info(f"OK: {site} (scope={scope}, version={version}, devices={len(target_device_ids)})")
            LOGGER.debug("Upgrade result for %s (%s): %s", site, sid, json.dumps(result))
            success += 1

        except Exception as e:
            console.error(f"FAIL: {site} (scope={scope}, version={version}) -> {e}")
            LOGGER.error("Upgrade failed for %s (%s): %s", site, sid, str(e))
            failed += 1

    console.info(f"\nDone. Success={success}, Skipped={skipped}, Failed={failed}. Log: {LOG_FILE}")
    
    # Exit with proper exit code
    if failed > 0:
        sys.exit(1)
    sys.exit(0)
